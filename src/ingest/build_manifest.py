"""
Build the IPEDS download manifest from NCES documentation.

Why this exists rather than a hardcoded list of filenames:

IPEDS table names follow at least three different conventions inside a single
collection year. From 2023-24:

    HD2023          single year label
    EF2023A         single year plus a part suffix
    C2023_A         underscore part suffix
    GR200_23        two digit year in the middle of the name
    SFA2223_P1      two year label, because aid spans an academic year
    F2223_F1A       two year label plus a form suffix

Any pattern generated from a template will therefore be wrong for some
surveys, and silently wrong, which is worse. NCES publishes a documentation
workbook for each collection year that lists every table it contains. This
script downloads those workbooks and reads the authoritative table list out
of them.

Output: config/ipeds_manifest.csv, one row per table per year, with the
survey it belongs to, its release status, and the direct download URL.

Run:
    python -m src.ingest.build_manifest
"""

from __future__ import annotations

import csv
import time
from pathlib import Path

import openpyxl
import requests

from src.common import PROJECT_ROOT, get_config, get_logger, resolve_path

log = get_logger("ingest.build_manifest")

TABLESDOC_URL = "https://nces.ed.gov/ipeds/tablefiles/tableDocs/IPEDS{label}Tablesdoc.xlsx"
DATA_FILE_URL = "https://nces.ed.gov/ipeds/datacenter/data/{table}.zip"

MANIFEST_PATH = PROJECT_ROOT / "config" / "ipeds_manifest.csv"


def collection_label(year: int) -> str:
    """
    Turn a starting year into the label NCES uses in the Tablesdoc filename.

    IPEDS labels a collection by the academic year it covers, so the 2023-24
    collection is '202324'. 2014 becomes '201415'.
    """
    return f"{year}{(year + 1) % 100:02d}"


def download_tablesdoc(year: int, destination: Path, timeout: int, retries: int,
                       backoff: int) -> Path | None:
    """
    Download one collection year's documentation workbook.

    Returns the local path, or None if NCES does not publish one for that
    year. A missing year is a real possibility and is logged rather than
    treated as a crash, because the older collections are less consistent.
    """
    label = collection_label(year)
    url = TABLESDOC_URL.format(label=label)
    target = destination / f"IPEDS{label}Tablesdoc.xlsx"

    if target.exists():
        log.info("Tablesdoc %s already present, skipping download", label)
        return target

    for attempt in range(1, retries + 1):
        try:
            response = requests.get(url, timeout=timeout)
        except requests.RequestException as error:
            log.warning("Attempt %d for %s failed: %s", attempt, label, error)
            time.sleep(backoff)
            continue

        if response.status_code == 200:
            target.write_bytes(response.content)
            log.info(
                "Downloaded Tablesdoc %s, %d bytes", label, len(response.content)
            )
            return target

        if response.status_code == 404:
            log.warning(
                "No Tablesdoc published for %s (HTTP 404 at %s)", label, url
            )
            return None

        log.warning(
            "Attempt %d for %s returned HTTP %d", attempt, label, response.status_code
        )
        time.sleep(backoff)

    log.error("Gave up on Tablesdoc %s after %d attempts", label, retries)
    return None


def find_sheet(workbook: openpyxl.Workbook, prefix: str) -> str | None:
    """
    Locate a sheet by name prefix.

    Sheet names carry the two digit year, so the tables sheet is 'Tables23'
    in the 2023-24 workbook and 'Tables14' in 2014-15. Matching on prefix
    survives that, and survives NCES changing the suffix convention.
    """
    for name in workbook.sheetnames:
        if name.lower().startswith(prefix.lower()):
            return name
    return None


def read_tables(path: Path, year: int) -> list[dict[str, str]]:
    """Extract the table inventory from one Tablesdoc workbook."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    sheet_name = find_sheet(workbook, "Tables")
    if sheet_name is None:
        log.error(
            "No sheet starting with 'Tables' in %s. Sheets present: %s",
            path.name,
            workbook.sheetnames,
        )
        return []

    rows = list(workbook[sheet_name].iter_rows(values_only=True))
    if not rows:
        log.error("Sheet %s in %s is empty", sheet_name, path.name)
        return []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]

    def column(name: str) -> int | None:
        return header.index(name) if name in header else None

    index_survey = column("Survey")
    index_table = column("TableName")
    index_title = column("TableTitle")
    index_release = column("Release")

    if index_table is None:
        log.error(
            "Sheet %s in %s has no TableName column. Header: %s",
            sheet_name,
            path.name,
            header,
        )
        return []

    records: list[dict[str, str]] = []
    for row in rows[1:]:
        table_name = row[index_table] if index_table < len(row) else None
        if table_name is None or str(table_name).strip() in ("", "None"):
            continue

        table_name = str(table_name).strip()

        def value(index: int | None) -> str:
            if index is None or index >= len(row) or row[index] is None:
                return ""
            return str(row[index]).strip()

        records.append(
            {
                "collection_year": str(year),
                "collection_label": collection_label(year),
                "survey": value(index_survey),
                "table_name": table_name,
                "table_title": value(index_title),
                "release": value(index_release),
                "url": DATA_FILE_URL.format(table=table_name),
            }
        )

    log.info("Year %s: %d tables listed in %s", year, len(records), sheet_name)
    return records


def main() -> None:
    config = get_config()
    ingest = config["ingest"]

    year_start = ingest["year_start"]
    year_end = ingest["year_end"]
    timeout = ingest.get("request_timeout_seconds", 120)
    retries = ingest.get("max_retries", 3)
    backoff = ingest.get("retry_backoff_seconds", 5)
    delay = ingest.get("delay_between_requests_seconds", 1)

    dictionaries = resolve_path("raw") / "_dictionaries"
    dictionaries.mkdir(parents=True, exist_ok=True)

    log.info(
        "Building manifest for collection years %s to %s", year_start, year_end
    )

    all_records: list[dict[str, str]] = []
    missing_years: list[int] = []

    for year in range(year_start, year_end + 1):
        path = download_tablesdoc(year, dictionaries, timeout, retries, backoff)
        if path is None:
            missing_years.append(year)
            continue
        all_records.extend(read_tables(path, year))
        time.sleep(delay)

    if not all_records:
        log.error("No tables found for any year. Manifest not written.")
        return

    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "collection_year",
        "collection_label",
        "survey",
        "table_name",
        "table_title",
        "release",
        "url",
    ]
    with open(MANIFEST_PATH, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(all_records)

    surveys = sorted({record["survey"] for record in all_records if record["survey"]})
    years_found = sorted({record["collection_year"] for record in all_records})

    log.info("=" * 70)
    log.info("Manifest written to %s", MANIFEST_PATH)
    log.info("Total table entries : %d", len(all_records))
    log.info("Collection years    : %d (%s)", len(years_found), ", ".join(years_found))
    log.info("Distinct surveys    : %d", len(surveys))
    for survey in surveys:
        count = sum(1 for record in all_records if record["survey"] == survey)
        log.info("    %-38s %3d tables", survey, count)
    if missing_years:
        log.warning(
            "No Tablesdoc published for these years: %s",
            ", ".join(str(year) for year in missing_years),
        )
    log.info("=" * 70)


if __name__ == "__main__":
    main()
