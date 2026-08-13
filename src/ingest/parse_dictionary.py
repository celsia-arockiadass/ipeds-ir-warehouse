"""
Parse the annual IPEDS Tablesdoc workbooks into a machine readable catalog.

Three outputs, all written to data/interim:

  variable_catalog.csv   one row per variable per table per collection year,
                         with its title, data type, and the name of its
                         imputation flag column if it has one

  value_labels.csv       one row per coded value per variable per year, so a
                         code like SECTOR = 1 can be resolved to its label
                         without hardcoding lookup tables anywhere

  variable_changes.csv   one row per variable per table family, recording the
                         first and last year it appears and whether it spans
                         all ten years. This is the documented answer to
                         "variable renaming between survey years", which is a
                         Layer 1 requirement.

Why the catalog exists rather than reading the workbooks at load time: the ten
workbooks together hold roughly 27,000 variable rows and 130,000 value rows.
Parsing them once into flat CSVs makes every downstream step, the staging
loader, the data dictionary in Phase 13, the code value lookups in the
dimensions, a cheap join instead of an Excel read.

Run:
    python -m src.ingest.parse_dictionary
"""

from __future__ import annotations

import csv
from collections import defaultdict
from pathlib import Path

import openpyxl

from src.common import get_config, get_logger, resolve_path

log = get_logger("ingest.parse_dictionary")


def collection_label(year: int) -> str:
    return f"{year}{(year + 1) % 100:02d}"


def find_sheet(workbook: openpyxl.Workbook, prefix: str) -> str | None:
    """
    Locate a sheet by name prefix.

    Sheet names carry the two digit year, so the variable sheet is
    'varTable23' in the 2023-24 workbook and 'varTable14' in 2014-15.
    """
    for name in workbook.sheetnames:
        if name.lower().startswith(prefix.lower()):
            return name
    return None


def sheet_rows(workbook: openpyxl.Workbook, prefix: str,
               source: str) -> tuple[list[str], list[tuple]]:
    """Return the header and data rows of a sheet located by prefix."""
    name = find_sheet(workbook, prefix)
    if name is None:
        log.warning("No sheet starting with %r in %s. Sheets: %s",
                    prefix, source, workbook.sheetnames)
        return [], []

    rows = list(workbook[name].iter_rows(values_only=True))
    if not rows:
        log.warning("Sheet %s in %s is empty", name, source)
        return [], []

    header = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
    return header, rows[1:]


def cell(row: tuple, header: list[str], column: str) -> str:
    """
    Read a named column from a row, case insensitively.

    Case insensitivity is not defensive programming, it is required. NCES
    changed the capitalization of their own metadata column names in the
    2023-24 workbook:

        2014 to 2022 :  varName   imputationvar   varTitle   Tablenumber
        2023-24      :  VarName   ImputationVar   VarTitle   TableNumber

    An exact match parser reads nine years of workbooks, finds every sheet,
    matches no columns, and reports success with zero rows. That is exactly
    what happened on the first run.

    Missing columns are also tolerated, because NCES has added metadata
    columns over the years and a 2014 workbook lacks some a 2023 one has.
    """
    lookup = {name.lower(): index for index, name in enumerate(header)}
    index = lookup.get(column.lower())
    if index is None or index >= len(row) or row[index] is None:
        return ""
    value = str(row[index]).strip()
    return "" if value.lower() == "none" else value


def parse_year(path: Path, year: int) -> tuple[list[dict], list[dict]]:
    """Extract variables and value labels from one collection year."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)

    variables: list[dict] = []
    header, rows = sheet_rows(workbook, "varTable", path.name)
    for row in rows:
        name = cell(row, header, "VarName")
        table = cell(row, header, "TableName")
        if not name or not table:
            continue
        variables.append(
            {
                "collection_year": year,
                "table_name": table,
                "survey": cell(row, header, "Survey"),
                "var_name": name,
                "var_title": cell(row, header, "VarTitle"),
                "data_type": cell(row, header, "DataType"),
                "field_width": cell(row, header, "FieldWidth"),
                "imputation_var": cell(row, header, "ImputationVar"),
                "long_description": cell(row, header, "LongDescription")[:500],
            }
        )

    labels: list[dict] = []
    header, rows = sheet_rows(workbook, "valueSets", path.name)
    for row in rows:
        name = cell(row, header, "VarName")
        table = cell(row, header, "TableName")
        code = cell(row, header, "Codevalue")
        if not name or not table or code == "":
            continue
        labels.append(
            {
                "collection_year": year,
                "table_name": table,
                "var_name": name,
                "code_value": code,
                "value_label": cell(row, header, "ValueLabel"),
                "frequency": cell(row, header, "Frequency"),
            }
        )

    # A sheet that exists, has rows, and yields nothing usable is the most
    # dangerous outcome available: it looks like success. Say so loudly.
    if not variables:
        log.error(
            "Year %d produced ZERO variables from %s. The sheet was found and "
            "was not empty, so the column names did not match. Check the "
            "header row.", year, path.name,
        )
    if not labels:
        log.error("Year %d produced ZERO value labels from %s", year, path.name)

    log.info("Year %d: %d variables, %d value labels", year, len(variables),
             len(labels))
    return variables, labels


def table_family(table_name: str) -> str:
    """
    Reduce a table name to a family that is stable across years, so the same
    logical table can be tracked over time.

    HD2023 becomes HD, EF2023A becomes EF*A, GR200_23 becomes GR200,
    SFA2223_P1 becomes SFA*_P1, F2223_F1A becomes F*_F1A.

    Deliberately crude. Its only job is grouping for the change report, and
    nothing downstream depends on it. Table names for actual work always come
    from the manifest.
    """
    import re

    name = re.sub(r"\d{4}", "*", table_name)
    name = re.sub(r"_\*$", "", name)
    name = re.sub(r"_\d{2}$", "", name)
    return name


def build_change_report(variables: list[dict]) -> list[dict]:
    """
    Record, for each variable in each table family, the years it appears in.

    A variable that stops appearing part way through the range, or starts
    late, is a discontinuity that any ten year trend has to account for. This
    is the evidence behind the limitations section rather than a guess.
    """
    appearances: dict[tuple[str, str], set[int]] = defaultdict(set)
    titles: dict[tuple[str, str], str] = {}

    for record in variables:
        key = (table_family(record["table_name"]), record["var_name"])
        appearances[key].add(record["collection_year"])
        titles.setdefault(key, record["var_title"])

    all_years = {record["collection_year"] for record in variables}
    total_years = len(all_years)

    report: list[dict] = []
    for (family, var_name), years in sorted(appearances.items()):
        ordered = sorted(years)
        report.append(
            {
                "table_family": family,
                "var_name": var_name,
                "var_title": titles[(family, var_name)],
                "first_year": ordered[0],
                "last_year": ordered[-1],
                "years_present": len(ordered),
                "spans_all_years": "yes" if len(ordered) == total_years else "no",
                "gap_in_middle": "yes"
                if len(ordered) != (ordered[-1] - ordered[0] + 1)
                else "no",
            }
        )
    return report


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    with open(path, "w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
    log.info("Wrote %-24s %6d rows", path.name, len(rows))


def main() -> None:
    config = get_config()
    year_start = config["ingest"]["year_start"]
    year_end = config["ingest"]["year_end"]

    dictionaries = resolve_path("raw") / "_dictionaries"
    interim = resolve_path("interim")

    all_variables: list[dict] = []
    all_labels: list[dict] = []

    for year in range(year_start, year_end + 1):
        path = dictionaries / f"IPEDS{collection_label(year)}Tablesdoc.xlsx"
        if not path.exists():
            log.warning("Missing workbook %s, skipping year %d", path.name, year)
            continue
        variables, labels = parse_year(path, year)
        all_variables.extend(variables)
        all_labels.extend(labels)

    if not all_variables:
        log.error("No variables parsed. Run: python -m src.ingest.build_manifest")
        return

    write_csv(
        interim / "variable_catalog.csv",
        all_variables,
        ["collection_year", "table_name", "survey", "var_name", "var_title",
         "data_type", "field_width", "imputation_var", "long_description"],
    )

    write_csv(
        interim / "value_labels.csv",
        all_labels,
        ["collection_year", "table_name", "var_name", "code_value",
         "value_label", "frequency"],
    )

    changes = build_change_report(all_variables)
    write_csv(
        interim / "variable_changes.csv",
        changes,
        ["table_family", "var_name", "var_title", "first_year", "last_year",
         "years_present", "spans_all_years", "gap_in_middle"],
    )

    flagged = sum(1 for v in all_variables if v["imputation_var"])
    unstable = [c for c in changes if c["spans_all_years"] == "no"]
    gapped = [c for c in changes if c["gap_in_middle"] == "yes"]

    years_parsed = sorted({v["collection_year"] for v in all_variables})
    expected = list(range(year_start, year_end + 1))

    log.info("=" * 70)
    log.info("Years parsed             : %d of %d", len(years_parsed),
             len(expected))
    if len(years_parsed) != len(expected):
        # Without this check the "present in every year" figure below is
        # computed over whatever subset happened to parse, which reads as a
        # clean result and is not one.
        missing = [y for y in expected if y not in years_parsed]
        log.error("MISSING YEARS: %s. Every figure below is computed over the "
                  "years that did parse and is not trustworthy until this is "
                  "fixed.", ", ".join(str(y) for y in missing))
    log.info("Variables catalogued     : %d", len(all_variables))
    log.info("  with an imputation flag: %d (%.1f%%)", flagged,
             100 * flagged / len(all_variables))
    log.info("Value labels catalogued  : %d", len(all_labels))
    log.info("Distinct variables       : %d", len(changes))
    log.info("  present in every year  : %d", len(changes) - len(unstable))
    log.info("  NOT in every year      : %d", len(unstable))
    log.info("  with a gap in the middle: %d", len(gapped))
    log.info("=" * 70)


if __name__ == "__main__":
    main()
